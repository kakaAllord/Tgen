"""Builds a small synthetic master workbook that mimics the real timetable
structure described in plan.md, for local smoke-testing of the pipeline.

Not part of the shipped application - dev/test tooling only.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

SAMPLE_PATH = Path(__file__).parent / "sample_master.xlsx"


def build_sample_workbook(path: Path = SAMPLE_PATH) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    _build_level_sheet(
        workbook,
        "Form 1",
        headers=["TIME/DAY", "07:30-08:10", "08:10-08:50", "08:50-09:30", "09:30-09:50", "09:50-10:30"],
        double_period_cols=(3, 4),  # merge "08:10-08:50" and "08:50-09:30" headers together
        rows={
            "MONDAY": ["PARADE", "EET- Mr. Mathias", None, "MATH-Mrs. Juma", "TEA BREAK", "CA&CAD Mr. Mbelwa"],
            "TUESDAY": ["TD-Mr. Marobo", None, None, "ENGLISH- Ms. Kway", "TEA BREAK", ""],
            "WEDNESDAY": ["MATH-Mrs. Juma", "EET- Mr. Mathias", None, None, "TEA BREAK", "TD-Mr. Marobo"],
            "THURSDAY": ["", "", "", "CA&CAD Mr. Mbelwa", "TEA BREAK", "EET- Mr. Mathias"],
            "FRIDAY": ["ENGLISH- Ms. Kway", "MATH-Mrs. Juma", None, "", "TEA BREAK", "GAMES"],
        },
        merge_cell=(2, 3),  # row 2 of data (TUESDAY), columns C:D (0-indexed 2,3) -> double period
    )

    _build_level_sheet(
        workbook,
        "Form 2",
        headers=["TIME/DAY", "07:30-08:10", "08:10-08:50", "08:50-09:30"],
        double_period_cols=None,
        rows={
            "MONDAY": ["CA&CAD Mr. Mbelwa", "TD-Mr. Marobo", "TEA BREAK"],
            "TUESDAY": ["EET- Mr. Mathias", "", "TEA BREAK"],
            "WEDNESDAY": ["", "CA&CAD Mr. Mbelwa", "TEA BREAK"],
        },
        merge_cell=None,
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _build_level_sheet(
    workbook: Workbook,
    title: str,
    headers: list[str],
    double_period_cols: tuple[int, int] | None,
    rows: dict[str, list[str | None]],
    merge_cell: tuple[int, int] | None,
) -> None:
    sheet = workbook.create_sheet(title=title)

    # A bit of noise above the header row, like a title banner in the real file.
    sheet.cell(row=1, column=1, value=f"{title.upper()} TIMETABLE - TERM 1")

    header_row_idx = 3
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=header_row_idx, column=col_idx, value=header)

    if double_period_cols is not None:
        start_col, end_col = double_period_cols
        sheet.merge_cells(
            start_row=header_row_idx, start_column=start_col, end_row=header_row_idx, end_column=end_col
        )

    for row_offset, (day, cells) in enumerate(rows.items()):
        row_idx = header_row_idx + 1 + row_offset
        sheet.cell(row=row_idx, column=1, value=day)
        for col_offset, value in enumerate(cells, start=2):
            sheet.cell(row=row_idx, column=col_offset, value=value)

    if merge_cell is not None:
        row_offset, start_col = merge_cell
        row_idx = header_row_idx + 1 + row_offset
        sheet.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=start_col + 1)


if __name__ == "__main__":
    output_path = build_sample_workbook()
    print(f"Sample workbook written to {output_path}")
