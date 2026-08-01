"""Exports per-teacher timetables to styled .xlsx files."""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import TeacherTimetable
from utils import CLASH_COLORS, INSTITUTION_NAME, fit_row_heights, sanitize_filename

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_DAY_FONT = Font(bold=True)
_BORDER = Border(*(Side(style="thin", color="B7B7B7") for _ in range(4)))
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _cell_value(entries: list[str]) -> str | CellRichText:
    """Plain text for a normal cell; for a double-booking (2+ entries),
    rich text with each lesson on its own line in a different color so the
    clash is unmistakable even on a black-and-white printout.
    """
    if not entries:
        return ""
    if len(entries) == 1:
        return entries[0]
    blocks: list[str | TextBlock] = []
    for i, entry in enumerate(entries):
        if i > 0:
            blocks.append("\n")
        color = CLASH_COLORS[i % len(CLASH_COLORS)]
        blocks.append(TextBlock(InlineFont(color=color, b=True), entry))
    return CellRichText(*blocks)

# Virtual A4-landscape page (in points) used to size rows so a timetable with
# few lessons still fills the printed page instead of leaving it half blank.
_PAGE_HEIGHT_PT = 595.0
_MARGIN_PT = 24.0
_INSTITUTION_ROW_HEIGHT = 26.0
_TITLE_ROW_HEIGHT = 22.0
_HEADER_ROW_HEIGHT = 22.0
_MIN_DATA_ROW_HEIGHT = 22.0
_MAX_DATA_ROW_HEIGHT = 70.0
_BOTTOM_PADDING = 16.0


def export_teacher_timetable(timetable: TeacherTimetable, output_dir: Path) -> Path:
    """Write one teacher's timetable to ``output_dir``, overwriting any existing file."""
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{sanitize_filename(timetable.teacher)}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Timetable"

    n_cols = len(timetable.days) + 1
    institution_row = 1
    title_row = 2
    header_row = 3

    sheet.merge_cells(
        start_row=institution_row, start_column=1, end_row=institution_row, end_column=max(n_cols, 1)
    )
    institution_cell = sheet.cell(row=institution_row, column=1, value=INSTITUTION_NAME)
    institution_cell.font = _TITLE_FONT
    institution_cell.alignment = _CENTER
    sheet.row_dimensions[institution_row].height = _INSTITUTION_ROW_HEIGHT

    sheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=max(n_cols, 1))
    title_cell = sheet.cell(row=title_row, column=1, value=f"{timetable.teacher} - Timetable")
    title_cell.font = _TITLE_FONT
    title_cell.alignment = _CENTER
    sheet.row_dimensions[title_row].height = _TITLE_ROW_HEIGHT

    sheet.cell(row=header_row, column=1, value="TIME / DAY")
    for col_idx, day in enumerate(timetable.days, start=2):
        sheet.cell(row=header_row, column=col_idx, value=day)
    for cell in sheet[header_row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
    sheet.row_dimensions[header_row].height = _HEADER_ROW_HEIGHT

    used_height = _INSTITUTION_ROW_HEIGHT + _TITLE_ROW_HEIGHT + _BOTTOM_PADDING
    available_table_height = _PAGE_HEIGHT_PT - 2 * _MARGIN_PT - used_height
    data_row_height = fit_row_heights(
        n_data_rows=len(timetable.time_slots),
        available_height=available_table_height,
        header_height=_HEADER_ROW_HEIGHT,
        min_data_row_height=_MIN_DATA_ROW_HEIGHT,
        max_data_row_height=_MAX_DATA_ROW_HEIGHT,
    )

    for row_idx, time_slot in enumerate(timetable.time_slots, start=header_row + 1):
        time_cell = sheet.cell(row=row_idx, column=1, value=time_slot)
        time_cell.font = _DAY_FONT
        time_cell.alignment = _LEFT
        time_cell.border = _BORDER

        activity_label = timetable.row_activities.get(time_slot)
        if activity_label:
            last_col = max(n_cols, 2)
            sheet.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=last_col)
            for col_idx in range(2, last_col + 1):
                sheet.cell(row=row_idx, column=col_idx).border = _BORDER
            activity_cell = sheet.cell(row=row_idx, column=2, value=activity_label.upper())
            activity_cell.font = _DAY_FONT
            activity_cell.alignment = _CENTER
        else:
            for col_idx, day in enumerate(timetable.days, start=2):
                entries = timetable.cell_entries(day, time_slot)
                cell = sheet.cell(row=row_idx, column=col_idx, value=_cell_value(entries))
                cell.alignment = _CENTER
                cell.border = _BORDER
        sheet.row_dimensions[row_idx].height = data_row_height

    sheet.freeze_panes = f"B{header_row + 1}"
    sheet.column_dimensions["A"].width = 16
    for col_idx in range(2, n_cols + 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 24

    last_row = header_row + len(timetable.time_slots)
    sheet.print_area = f"A1:{get_column_letter(max(n_cols, 1))}{max(last_row, header_row)}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    workbook.save(path)
    logger.debug("Wrote Excel timetable for %s -> %s", timetable.teacher, path)
    return path


def export_all(timetables: dict[str, TeacherTimetable], output_dir: Path) -> list[Path]:
    """Export every teacher's timetable to ``output_dir``. Returns the written paths."""
    paths = [export_teacher_timetable(timetable, output_dir) for timetable in timetables.values()]
    logger.info("Exported %d Excel timetable(s) to %s", len(paths), output_dir)
    return paths
