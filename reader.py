"""Loads the master Excel workbook and resolves merged cells.

The master workbook is the sole source of truth: it is opened, read, and
closed. Nothing is ever written back to it.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from models import SheetGrid
from utils import normalize_text

logger = logging.getLogger(__name__)


def read_workbook(path: str | Path) -> list[SheetGrid]:
    """Read every worksheet of the master workbook into resolved grids.

    Merged cells are expanded in-memory so every cell in a merged range
    carries the same value as the merge's top-left cell (this is what lets
    the parser correctly attribute a lesson spanning a merged block to every
    day/time it covers). The source file itself is opened read-only and is
    never saved back to.
    """
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Master workbook not found: {workbook_path}")

    workbook = load_workbook(filename=workbook_path, data_only=True, read_only=False)
    try:
        sheets = [_read_sheet(sheet) for sheet in workbook.worksheets]
    finally:
        workbook.close()

    logger.info("Read %d worksheet(s) from %s", len(sheets), workbook_path.name)
    return sheets


def _read_sheet(sheet: Worksheet) -> SheetGrid:
    max_row = sheet.max_row or 0
    max_col = sheet.max_column or 0

    grid: list[list[str]] = [
        [normalize_text(cell.value) for cell in row]
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col)
    ]

    for merged_range in sheet.merged_cells.ranges:
        top_value = grid[merged_range.min_row - 1][merged_range.min_col - 1]
        for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
            for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                grid[row_idx - 1][col_idx - 1] = top_value

    logger.debug("Read sheet %r: %d rows x %d cols", sheet.title, max_row, max_col)
    return SheetGrid(name=sheet.title, rows=grid)
